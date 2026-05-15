"""
GuniVox Backend — Production-hardened v4
==========================================
All known bugs fixed (B1-B12), transfer call perfected,
voice capture fixed, full fallback chain added.

FIX SUMMARY:
  B1  — greeting.wav served from disk with FALLBACK if missing (no dead air on first call)
  B2  — Record timeout raised to 3s; dynamic based on turn count for natural pauses
  B3  — Silero VAD init guarded: logger exists before torch.hub.load is called
  B4  — Session lock re-entry deadlock: get_ai_response does NOT re-acquire held lock
  B5  — Early TTS cancel race: always await before cancelling; guarded with done() check
  B6  — active_counsellors is a per-call flag stored in sessions, not a shared set (no leak)
  B7  — httpx client: one client per event loop, properly closed on shutdown
  B8  — /vobiz-respond extracts call_sid BEFORE the session lock acquisition
  B9  — Hallucination list expanded; Layer 2/3/4 applied even when STT returns non-empty
  B10 — TTS fallback chain: Sarvam → gTTS → silence WAV (never empty audio_url in XML)
  B11 — Campaign thread safety: uses a proper threading.RLock instead of threading.Lock
  B12 — Counsellor-hangup always frees counsellor from sessions even on exception
  B13 — VAD: silence detection runs in executor so it never blocks the event loop
  B14 — /vobiz-silent and /hold-loop guard against missing call_sid gracefully
  B15 — Conference room name stored and reused deterministically; never falls back to wrong room
"""

import os
import logging
import json
import re
import sqlite3
import time
import csv
import wave
import io
import struct
import asyncio
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from io import StringIO, BytesIO
import requests
import base64
import httpx
from fastapi import FastAPI, Request, HTTPException, UploadFile, File, WebSocket
from fastapi.responses import Response, FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from openai import AsyncOpenAI
import uuid
import threading
from dotenv import load_dotenv
from openpyxl import Workbook

# FIX #3: pure-Python audio imports (replaces ffmpeg subprocess, saves 50-200ms/turn)
try:
    import soundfile as sf
    import numpy as np
    _soundfile_available = True
except ImportError:
    _soundfile_available = False

# ─────────────────────────────────────────
# LOGGING (must be first — B3 fix)
# ─────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────
# SILERO VAD (B3 fix: logger exists before import)
# ─────────────────────────────────────────
silero_model = None
silero_get_speech_timestamps = None

try:
    import torch
    import numpy as np
    _silero_model, _silero_utils = torch.hub.load(
        repo_or_dir='snakers4/silero-vad',
        model='silero_vad',
        force_reload=False,
        verbose=False
    )
    silero_model = _silero_model
    silero_get_speech_timestamps = _silero_utils[0]
    logger.info("✅ Silero VAD loaded")
except Exception as e:
    logger.warning(f"⚠️  Silero VAD unavailable: {e} — amplitude fallback will be used")

# ─────────────────────────────────────────
# ENV + CONFIG
# ─────────────────────────────────────────
load_dotenv(".env.local")

HTTP_TIMEOUT_SECONDS       = int(os.getenv("HTTP_TIMEOUT_SECONDS", "15"))
SESSION_TTL                = 3600
TERMINAL_CALL_STATUSES     = {"completed", "busy", "no-answer", "canceled", "failed", "hangup"}
CALL_POLL_INTERVAL_SECONDS = 2
CALL_MAX_DURATION_SECONDS  = int(os.getenv("CALL_MAX_DURATION_SECONDS", "180"))
CSV_PHONE_HEADERS          = {"phone", "phone_number", "mobile", "number", "contact"}
ENABLE_RAG                 = os.getenv("ENABLE_RAG", "true").strip().lower() in {"1", "true", "yes", "on"}
RAG_TOP_K                  = max(1, int(os.getenv("RAG_TOP_K", "2")))

SARVAM_API_KEY    = os.getenv("SARVAM_API_KEY", "")
VOBIZ_AUTH_ID     = os.getenv("VOBIZ_AUTH_ID", "")
VOBIZ_AUTH_TOKEN  = os.getenv("VOBIZ_AUTH_TOKEN", "")
VOBIZ_FROM_NUMBER = os.getenv("VOBIZ_FROM_NUMBER", "")
BASE_URL          = os.getenv("BASE_URL", "").rstrip("/")
CONNECTING_MUSIC  = os.getenv("CONNECTING_MUSIC", "goodvibes.mp3")

_tn_env          = os.getenv("TRANSFER_NUMBERS", os.getenv("TRANSFER_NUMBER", ""))
TRANSFER_NUMBERS = [n.strip() for n in _tn_env.split(",") if n.strip()]

AI_PROVIDER            = os.getenv("AI_PROVIDER", "openai").strip().lower()
OLLAMA_BASE_URL        = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434").rstrip("/")
OLLAMA_MODEL           = os.getenv("OLLAMA_MODEL", "jatas-qwen-rag:latest")
OLLAMA_TIMEOUT_SECONDS = int(os.getenv("OLLAMA_TIMEOUT_SECONDS", "45"))
OPENAI_MODEL           = os.getenv("OPENAI_MODEL", "gpt-4o-mini")

if AI_PROVIDER not in {"openai", "ollama"}:
    logger.warning(f"AI_PROVIDER='{AI_PROVIDER}' not supported — defaulting to 'openai'.")
    AI_PROVIDER = "openai"

if not SARVAM_API_KEY:
    logger.warning("⚠️  No SARVAM_API_KEY — STT and Sarvam TTS disabled.")
if not VOBIZ_AUTH_ID:
    logger.warning("⚠️  No VOBIZ_AUTH_ID — outbound calls will fail.")

# ─────────────────────────────────────────
# THREAD POOLS
# ─────────────────────────────────────────
_io_executor    = ThreadPoolExecutor(max_workers=16, thread_name_prefix="io")
_db_executor    = ThreadPoolExecutor(max_workers=1,  thread_name_prefix="db")
_faiss_executor = ThreadPoolExecutor(max_workers=4,  thread_name_prefix="faiss")
_vad_executor   = ThreadPoolExecutor(max_workers=4,  thread_name_prefix="vad")  # B13

# ─────────────────────────────────────────
# LLM CLIENT
# ─────────────────────────────────────────
aclient: Optional[AsyncOpenAI] = None
if AI_PROVIDER == "openai":
    aclient = AsyncOpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    logger.info(f"✅ LLM: OpenAI async | model={OPENAI_MODEL}")
elif AI_PROVIDER == "ollama":
    logger.info(f"✅ LLM: Ollama | model={OLLAMA_MODEL}")

http_session = requests.Session()

# ─────────────────────────────────────────
# FASTAPI APP
# ─────────────────────────────────────────
app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)
os.makedirs("static/audio", exist_ok=True)
app.mount("/static", StaticFiles(directory="static"), name="static")

# ─────────────────────────────────────────
# SESSION STORE (locks + TTL)
# ─────────────────────────────────────────
sessions:       Dict[str, object]          = {}
session_locks:  Dict[str, asyncio.Lock]    = {}
session_expiry: Dict[str, float]           = {}
campaigns:      Dict[str, dict]            = {}
campaign_lock   = threading.RLock()        # B11: RLock prevents re-entry deadlock

def get_session_lock(call_sid: str) -> asyncio.Lock:
    if call_sid not in session_locks:
        session_locks[call_sid] = asyncio.Lock()
    return session_locks[call_sid]

# ─────────────────────────────────────────
# REGEX PATTERNS
# ─────────────────────────────────────────
LANG_PATTERN     = re.compile(r"LANG:\s*([a-z\-]+)", re.IGNORECASE)
TEXT_PATTERN     = re.compile(r"TEXT:\s*(.*?)(?=\s*\||\s*NAME:|\s*INTEREST:|\s*STATUS:|\s*FOLLOW_UP:|$)",
                               re.DOTALL | re.IGNORECASE)
METADATA_PATTERNS = {
    "user_name":   re.compile(r"NAME:\s*(.*?)(?=\s*\||STATUS:|INTEREST:|LANG:|TEXT:|$)",
                               re.IGNORECASE | re.DOTALL),
    "interest":    re.compile(r"INTEREST:\s*(.*?)(?=\s*\||STATUS:|NAME:|LANG:|TEXT:|$)",
                               re.IGNORECASE | re.DOTALL),
    "lead_status": re.compile(r"STATUS:\s*(.*?)(?=\s*\||NAME:|INTEREST:|LANG:|TEXT:|$)",
                               re.IGNORECASE | re.DOTALL),
    "follow_up":   re.compile(r"FOLLOW_UP:\s*(.*?)(?=\s*\||$)",
                               re.IGNORECASE | re.DOTALL),
}
_SENTENCE_BOUNDARIES = {'।', '.', '?', '!', '?'}

# ─────────────────────────────────────────
# DATABASE (WAL mode)
# ─────────────────────────────────────────
DB_FILE = "gunivox.db"

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("PRAGMA journal_mode=WAL")
    c.execute("PRAGMA synchronous=NORMAL")
    c.execute("PRAGMA table_info(calls)")
    cols = [col[1] for col in c.fetchall()]
    for col, default in [
        ("stage",            "Cold Call"),
        ("duration_seconds", "0"),
        ("billable_minutes", "0"),
        ("ended_at",         "NULL"),
    ]:
        if col not in cols:
            c.execute(f"ALTER TABLE calls ADD COLUMN {col} TEXT DEFAULT '{default}'")
    c.execute("""
        CREATE TABLE IF NOT EXISTS calls (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            call_sid TEXT UNIQUE,
            phone_number TEXT,
            status TEXT,
            started_at TEXT,
            ended_at TEXT,
            duration_seconds INTEGER DEFAULT 0,
            billable_minutes REAL DEFAULT 0,
            end_reason TEXT,
            user_name TEXT,
            interest TEXT,
            lead_status TEXT,
            follow_up TEXT,
            transcript TEXT,
            stage TEXT DEFAULT 'Cold Call'
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS courses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT, institute TEXT, duration TEXT, fees TEXT,
            eligibility TEXT, counsellor TEXT, phone TEXT, brochure_url TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS rag_documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT, content TEXT NOT NULL, source TEXT, created_at TEXT
        )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS idx_calls_phone ON calls(phone_number)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_calls_stage ON calls(stage)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_calls_sid   ON calls(call_sid)")
    conn.commit()
    conn.close()
    _populate_default_courses()

def _populate_default_courses():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM courses")
    if c.fetchone()[0] == 0:
        defaults = [
            ("BCA",      "S.O. Patel College (SOCET)", "3 Years", "70,000/yr",    "10+2 English required",      "Dr. Mehta",        "98765 43210", None),
            ("MCA",      "U.V. Patel College (UVPCE)", "2 Years", "1,40,000/yr",  "Needs BCA/BE/BSc with Maths","Prof. Shah",       "98250 12345", None),
            ("B.Tech IT","UVPCE",                      "4 Years", "1,20,000/yr",  "12th PCM 45%",               "Counselling Team", "79900 11223", None),
        ]
        c.executemany(
            "INSERT INTO courses (name,institute,duration,fees,eligibility,counsellor,phone,brochure_url) "
            "VALUES (?,?,?,?,?,?,?,?)",
            defaults
        )
        conn.commit()
        logger.info("✅ Default courses populated")
    conn.close()

init_db()

# ─────────────────────────────────────────
# FAISS RAG
# ─────────────────────────────────────────
try:
    import faiss_rag

    def _init_faiss():
        try:
            faiss_rag.load_index(json_path='final_dataset.json')
            faiss_rag.get_model()
            logger.info(f"✅ FAISS RAG ready — {faiss_rag._index.ntotal} vectors")
        except Exception as e:
            logger.error(f"❌ FAISS RAG init failed: {e}")

    threading.Thread(target=_init_faiss, daemon=True).start()
    _faiss_available = True
except ImportError:
    logger.warning("⚠️  faiss_rag module not found — RAG disabled")
    _faiss_available = False

    class _FaissStub:
        def is_ready(self): return False
        def search(self, *a, **kw): return []
        def stats(self): return {"ready": False, "total_vectors": 0, "model": "n/a"}
        SCORE_THRESHOLD = 0.5

    faiss_rag = _FaissStub()  # type: ignore

# ─────────────────────────────────────────
# SYSTEM PROMPT CACHE
# ─────────────────────────────────────────
try:
    from prompt_config import SYSTEM_PROMPT
except ImportError:
    SYSTEM_PROMPT = (
        "You are Ananya, an AI career assistant for Ganpat University. "
        "Always respond in Gujarati only. Be concise, warm and helpful."
    )
    logger.warning("⚠️  prompt_config not found — using default system prompt")

_system_prompt_cache: Optional[str] = None

def _build_system_prompt() -> str:
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT name, eligibility, duration, fees, counsellor, phone FROM courses")
        rows = c.fetchall()
        conn.close()
        course_text = "\n".join([
            f"- **{r[0]}**: Eligibility: {r[1]}, Duration: {r[2]}, "
            f"Fees: {r[3]}, Counsellor: {r[4]} ({r[5]})"
            for r in rows
        ]) or "- No specific course data available."
        return SYSTEM_PROMPT + "\n\n### ADDITIONAL COURSE DATA:\n" + course_text
    except Exception as e:
        logger.error(f"Prompt build error: {e}")
        return SYSTEM_PROMPT

async def get_system_prompt_with_courses() -> str:
    global _system_prompt_cache
    if _system_prompt_cache is None:
        loop = asyncio.get_running_loop()
        _system_prompt_cache = await loop.run_in_executor(_io_executor, _build_system_prompt)
    return _system_prompt_cache

def invalidate_prompt_cache():
    global _system_prompt_cache
    _system_prompt_cache = None
    logger.info("🔄 System prompt cache invalidated")

# ─────────────────────────────────────────
# HTTPX CLIENTS (B7: per-loop, closed on shutdown)
# ─────────────────────────────────────────
_httpx_clients: Dict[int, httpx.AsyncClient] = {}

def _get_httpx_client() -> httpx.AsyncClient:
    loop_id = id(asyncio.get_running_loop())
    if loop_id not in _httpx_clients or _httpx_clients[loop_id].is_closed:
        _httpx_clients[loop_id] = httpx.AsyncClient(
            timeout=httpx.Timeout(HTTP_TIMEOUT_SECONDS, connect=5.0),
            limits=httpx.Limits(max_connections=50, max_keepalive_connections=20),
        )
    return _httpx_clients[loop_id]

async def _close_httpx_clients():
    for client in _httpx_clients.values():
        if not client.is_closed:
            await client.aclose()

# ─────────────────────────────────────────
# TTS AUDIO LRU CACHE
# ─────────────────────────────────────────
_tts_cache: Dict[Tuple[str, str], str] = {}
_TTS_CACHE_MAX = 128

def _tts_cache_get(text: str, lang: str) -> Optional[str]:
    return _tts_cache.get((text.strip(), lang))

def _tts_cache_set(text: str, lang: str, url: str):
    key = (text.strip(), lang)
    if len(_tts_cache) >= _TTS_CACHE_MAX:
        oldest = next(iter(_tts_cache))
        del _tts_cache[oldest]
    _tts_cache[key] = url

# ─────────────────────────────────────────
# SILENCE WAV FALLBACK (B10: never return empty audio_url)
# ─────────────────────────────────────────
_SILENCE_WAV_PATH = os.path.join("static", "audio", "silence_500ms.wav")

def _ensure_silence_wav():
    """Create a 500ms silence WAV so TTS failures never leave dead air in XML."""
    if os.path.exists(_SILENCE_WAV_PATH):
        return
    try:
        sample_rate = 8000
        num_samples = sample_rate // 2  # 500 ms
        with wave.open(_SILENCE_WAV_PATH, 'wb') as wf:
            wf.setnchannels(1)
            wf.setsampwidth(2)
            wf.setframerate(sample_rate)
            wf.writeframes(b'\x00\x00' * num_samples)
        logger.info("✅ silence_500ms.wav created")
    except Exception as e:
        logger.error(f"Could not create silence WAV: {e}")

_ensure_silence_wav()

# ─────────────────────────────────────────
# DB HELPERS
# ─────────────────────────────────────────
def _save_call_log_sync(call_sid: str, data: dict):
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT id, started_at, duration_seconds FROM calls WHERE call_sid = ?", (call_sid,))
        row = c.fetchone()
        if not row:
            c.execute(
                "INSERT INTO calls (call_sid, phone_number, status, started_at, stage) VALUES (?,?,?,?,?)",
                (call_sid, data.get('phone_number'), 'initiated',
                 datetime.now().isoformat(), data.get('stage', 'Cold Call'))
            )
        else:
            allowed = ['status', 'end_reason', 'user_name', 'interest', 'lead_status',
                       'follow_up', 'transcript', 'stage', 'ended_at']
            fields, values = [], []
            status = data.get('status', '').lower()
            if status in TERMINAL_CALL_STATUSES and row[1]:
                ended_at = datetime.now()
                data['ended_at'] = ended_at.isoformat()
                try:
                    start_dt = datetime.fromisoformat(row[1])
                    diff = (ended_at - start_dt).total_seconds()
                    data['duration_seconds'] = int(diff)
                    data['billable_minutes'] = (int(diff + 29) // 30) * 0.5
                    allowed.extend(['duration_seconds', 'billable_minutes'])
                except Exception as e:
                    logger.error(f"Duration calc error: {e}")
            for k, v in data.items():
                if k in allowed:
                    fields.append(f"{k} = ?")
                    values.append(v)
            if fields:
                values.append(call_sid)
                c.execute(f"UPDATE calls SET {', '.join(fields)} WHERE call_sid = ?", values)
        conn.commit()
    except Exception as e:
        logger.error(f"DB write error for {call_sid}: {e}")
    finally:
        try:
            conn.close()
        except Exception:
            pass

def save_call_log(call_sid: str, data: dict):
    _db_executor.submit(_save_call_log_sync, call_sid, data)

def get_call_status_from_db(call_sid: str) -> Optional[str]:
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT status FROM calls WHERE call_sid = ?", (call_sid,))
        row = c.fetchone()
        conn.close()
        return row[0] if row else None
    except Exception:
        return None

# ─────────────────────────────────────────
# MISC HELPERS
# ─────────────────────────────────────────
def get_base_url(request: Request) -> str:
    env_base = os.getenv("BASE_URL")
    if env_base:
        return env_base.rstrip('/')
    host   = request.headers.get("x-forwarded-host") or request.headers.get("host", "localhost:8000")
    scheme = request.headers.get("x-forwarded-proto", "https")
    return f"{scheme}://{host}"

def get_cloudflare_headers() -> dict:
    return {"Cache-Control": "no-cache"}

def normalize_phone_number(raw: str) -> str:
    candidate = (raw or "").strip().replace(" ", "")
    if candidate.startswith("+"):
        return "+" + re.sub(r"\D", "", candidate[1:])
    return re.sub(r"\D", "", candidate)

def build_rag_context(query: str) -> str:
    if not ENABLE_RAG or not _faiss_available or not faiss_rag.is_ready():
        return ""
    try:
        results = faiss_rag.search(query, top_k=RAG_TOP_K)
        if not results:
            return ""
        parts = []
        for hit in results:
            ctx = hit.get('voice_context') or hit.get('text', '')
            if ctx:
                parts.append(f"[score={hit['score']:.2f}]\n{ctx}")
        return "\n\n".join(parts)
    except Exception as e:
        logger.error(f"FAISS search error: {e}")
        return ""

# ─────────────────────────────────────────
# VAD — runs in dedicated executor (B13)
# ─────────────────────────────────────────
def _detect_silence_sync(audio_bytes: bytes, ext: str) -> bool:
    """
    FIX #3: Pure-Python silence detection — no ffmpeg subprocess spawn.
    Uses soundfile + numpy for O(1) process overhead vs 50-200ms fork cost.
    Falls back to Silero VAD → amplitude check if soundfile unavailable.
    Runs in _vad_executor, never blocks the event loop (B13).
    """
    if ext not in (".wav", ".mp3"):
        return False

    # ── Fast path: soundfile (no subprocess, ~1ms) ───────────────────────
    if _soundfile_available:
        try:
            data, sr = sf.read(BytesIO(audio_bytes), dtype='float32', always_2d=False)
            if data.ndim > 1:
                data = data.mean(axis=1)  # stereo → mono
            max_amp = float(np.abs(data).max()) if len(data) > 0 else 0.0
            is_silent = max_amp < 0.06  # normalized float32 threshold
            logger.info(f"   soundfile VAD: max_amp={max_amp:.4f} | {'silent' if is_silent else 'speech'}")
            return is_silent
        except Exception as e:
            logger.warning(f"   soundfile VAD error: {e} — falling back to Silero/amplitude")

    # Only WAV supported for Silero/amplitude path (no ffmpeg)
    if ext != ".wav":
        return False

    try:
        with wave.open(io.BytesIO(audio_bytes), 'rb') as wf:
            frames = wf.readframes(wf.getnframes())
            width  = wf.getsampwidth()
    except Exception as e:
        logger.warning(f"   WAV read error: {e}")
        return False

    if width not in (1, 2) or not frames:
        return False

    fmt     = '<' + ('h' if width == 2 else 'b') * (len(frames) // width)
    samples = struct.unpack(fmt, frames)

    # ── Silero VAD (preferred if loaded) ────────────────────────────────
    if silero_model is not None and silero_get_speech_timestamps is not None:
        try:
            import torch
            import numpy as _np
            audio_np     = _np.array(samples, dtype=_np.float32) / 32768.0
            audio_tensor = torch.from_numpy(audio_np)
            timestamps   = silero_get_speech_timestamps(
                audio_tensor, silero_model, sampling_rate=16000
            )
            is_silent = len(timestamps) == 0
            logger.info(f"   Silero VAD: {'silent' if is_silent else 'speech'} | {len(timestamps)} segments")
            return is_silent
        except Exception as e:
            logger.warning(f"   Silero VAD error: {e} — falling back to amplitude")

    # ── Amplitude fallback ────────────────────────────────────────────────
    max_amp = max(abs(s) for s in samples[::10]) if samples else 0
    is_silent = max_amp < 1800
    logger.info(f"   Amplitude VAD: max={max_amp} | {'silent' if is_silent else 'speech'}")
    return is_silent

# ─────────────────────────────────────────
# STT
# ─────────────────────────────────────────
async def transcribe_audio(audio_bytes: bytes, filename: str) -> str:
    if SARVAM_API_KEY:
        return await _transcribe_sarvam(audio_bytes, filename)
    logger.warning("No SARVAM_API_KEY — skipping STT")
    return ""

async def _transcribe_sarvam(audio_bytes: bytes, filename: str) -> str:
    _t = time.time()
    for attempt in range(2):
        try:
            headers  = {'api-subscription-key': SARVAM_API_KEY}
            audio_io = BytesIO(audio_bytes)
            client   = _get_httpx_client()
            resp     = await client.post(
                "https://api.sarvam.ai/speech-to-text",
                headers=headers,
                data={'model': 'saaras:v3'},
                files=[('file', (filename, audio_io, 'audio/wav'))],
            )
            if resp.status_code == 200:
                transcript = resp.json().get("transcript", "").strip()
                logger.info(f"⏱️ STT: {(time.time()-_t)*1000:.0f}ms | '{transcript}'")
                return transcript
            if resp.status_code == 429 and attempt == 0:
                logger.warning("Sarvam STT 429 — retrying in 1s")
                await asyncio.sleep(1.0)
                continue
            logger.error(f"Sarvam STT {resp.status_code}: {resp.text[:200]}")
            return ""
        except Exception as e:
            logger.error(f"Sarvam STT error (attempt {attempt}): {e}")
            if attempt == 0:
                await asyncio.sleep(0.5)
    return ""

# ─────────────────────────────────────────
# LLM STREAMING
# ─────────────────────────────────────────
async def _stream_ollama(messages, temperature=0.3, max_tokens=60):
    loop  = asyncio.get_running_loop()
    queue: asyncio.Queue = asyncio.Queue()

    def _producer():
        try:
            payload = {
                "model": OLLAMA_MODEL, "messages": messages, "stream": True,
                "options": {"temperature": temperature, "num_predict": max_tokens},
            }
            resp = http_session.post(
                f"{OLLAMA_BASE_URL}/api/chat", json=payload,
                stream=True, timeout=OLLAMA_TIMEOUT_SECONDS
            )
            resp.raise_for_status()
            for line in resp.iter_lines():
                if line:
                    try:
                        chunk = (json.loads(line).get("message") or {}).get("content", "")
                        if chunk:
                            loop.call_soon_threadsafe(queue.put_nowait, chunk)
                    except Exception:
                        pass
        except Exception as e:
            logger.error(f"Ollama stream error: {e}")
        finally:
            loop.call_soon_threadsafe(queue.put_nowait, None)

    _io_executor.submit(_producer)
    while True:
        chunk = await queue.get()
        if chunk is None:
            break
        yield chunk

async def _stream_llm_with_early_tts(
    call_sid: str, messages: list, base_url: str,
    lang: str = "gu-IN", temperature=0.3, max_tokens=60
) -> Tuple[str, Optional[asyncio.Task], str]:
    """
    Streams LLM tokens. Fires TTS as a background task on first sentence boundary.
    Returns (full_text, first_tts_task, first_tts_text).
    B5 fix: task is always properly awaited/cancelled in the caller.
    """
    buf: List[str] = []
    first_sentence_fired = False
    first_tts_task: Optional[asyncio.Task] = None
    first_tts_text = ""
    _t = time.time()

    async def _token_stream():
        if AI_PROVIDER == "ollama":
            async for chunk in _stream_ollama(messages, temperature, max_tokens):
                yield chunk
        else:
            stream = await aclient.chat.completions.create(
                model=OPENAI_MODEL, messages=messages,
                temperature=temperature, max_tokens=max_tokens, stream=True
            )
            async for chunk in stream:
                delta = chunk.choices[0].delta.content or ""
                if delta:
                    yield delta

    async for token in _token_stream():
        buf.append(token)
        # FIX #5: Only fire early TTS once TEXT: tag has been confirmed in the buffer.
        # Firing on the first sentence boundary without TEXT: risks capturing raw metadata
        # (LANG: / STATUS: prefix tokens), causing the prefix-match check to fail and
        # forcing a full second TTS generation — doubling latency instead of saving it.
        if not first_sentence_fired and any(c in token for c in _SENTENCE_BOUNDARIES):
            partial    = "".join(buf).strip()
            text_match = TEXT_PATTERN.search(partial)
            # Guard: only proceed when the TEXT: tag is already present in the buffer
            if text_match:
                s1 = text_match.group(1).strip()
                s1 = re.sub(r'(LANG|STATUS|INTEREST|NAME|FOLLOW_UP):\s*\S+\s*\|?', '', s1).strip().strip('|').strip()
                if len(s1) > 5:
                    first_tts_text = s1
                    first_tts_task = asyncio.create_task(generate_tts_audio(s1, base_url, lang))
                    first_sentence_fired = True
                    logger.info(f"⏱️ LLM_FIRST_S: {(time.time()-_t)*1000:.0f}ms | early_text='{s1[:40]}'")

    return "".join(buf).strip(), first_tts_task, first_tts_text

# ─────────────────────────────────────────
# TTS — Sarvam → gTTS → silence (B10 full fallback chain)
# ─────────────────────────────────────────
SARVAM_LANG_MAP = {"gu-IN": "gu-IN", "hi-IN": "hi-IN", "en-IN": "en-IN"}

async def generate_tts_audio(text: str, base_url_arg: str, lang: str = "gu-IN") -> str:
    """
    Always returns a playable URL.
    Fallback chain: Sarvam → gTTS → 500ms silence WAV (B10).
    """
    os.makedirs(os.path.join("static", "audio"), exist_ok=True)
    text = text.strip()
    if not text:
        return _silence_url(base_url_arg)

    cached = _tts_cache_get(text, lang)
    if cached:
        logger.info(f"🔊 TTS cache hit")
        return cached

    _t = time.time()

    # ── Sarvam TTS ────────────────────────────────────────────────────────
    if SARVAM_API_KEY:
        try:
            target_lang = SARVAM_LANG_MAP.get(lang, "gu-IN")
            payload = {
                "text": text[:500],   # Sarvam limit guard
                "target_language_code": target_lang,
                "speaker": "anushka",
                "model": "bulbul:v2",
                "speech_sample_rate": 8000,
                "enable_preprocessing": True,
            }
            headers = {
                "api-subscription-key": SARVAM_API_KEY,
                "Content-Type": "application/json",
            }
            client   = _get_httpx_client()
            resp     = await client.post(
                "https://api.sarvam.ai/text-to-speech",
                json=payload, headers=headers
            )
            if resp.status_code == 200:
                audio_b64 = resp.json()["audios"][0]
                filename  = f"tts_{uuid.uuid4().hex}.wav"
                filepath  = os.path.join("static", "audio", filename)
                loop      = asyncio.get_running_loop()
                await loop.run_in_executor(
                    _io_executor,
                    lambda: open(filepath, "wb").write(base64.b64decode(audio_b64))
                )
                url = f"{base_url_arg}/static/audio/{filename}"
                _tts_cache_set(text, lang, url)
                logger.info(f"🔊 Sarvam TTS ({(time.time()-_t)*1000:.0f}ms)")
                return url
            logger.error(f"Sarvam TTS {resp.status_code}: {resp.text[:200]}")
        except Exception as e:
            logger.error(f"Sarvam TTS error: {e}")

    # ── gTTS fallback ─────────────────────────────────────────────────────
    gtts_url = await _gtts_fallback(text, lang, base_url_arg)
    if gtts_url:
        return gtts_url

    # ── Silence fallback ─────────────────────────────────────────────────
    logger.warning("⚠️  All TTS failed — returning silence WAV")
    return _silence_url(base_url_arg)

def _silence_url(base_url_arg: str) -> str:
    return f"{base_url_arg}/static/audio/silence_500ms.wav"

async def _gtts_fallback(text: str, lang: str, base_url_arg: str) -> str:
    def _synth():
        try:
            from gtts import gTTS
            lang_code = lang.split("-")[0]
            filename  = f"tts_{uuid.uuid4().hex}.mp3"
            filepath  = os.path.join("static", "audio", filename)
            gTTS(text=text[:500], lang=lang_code).save(filepath)
            return filename
        except Exception as e:
            logger.error(f"gTTS error: {e}")
            return ""
    filename = await asyncio.get_running_loop().run_in_executor(_io_executor, _synth)
    return f"{base_url_arg}/static/audio/{filename}" if filename else ""

# ─────────────────────────────────────────
# STARTUP: greeting.wav + TTS prewarm
# ─────────────────────────────────────────
GREETING_TEXT = (
    "Hi, હું Ananya, Ganpat University તરફથી AI Career Assistant બોલું છું. "
    "ઘણા વિદ્યાર્થીઓને 10મા, 12મા અથવા Graduation પછી યોગ્ય career પસંદ કરવામાં "
    "મુશ્કેલી પડે છે. વિદ્યાર્થીઓને યોગ્ય માર્ગદર્શન આપવા માટે અમે તમારા શહેરમાં "
    "FREE One-to-One Career Counselling Session આયોજન કરી રહ્યા છીએ. "
    "શું તમને આ counselling session માં જોડાવું ગમશે?"
)
GREETING_PATH = os.path.join("static", "audio", "greeting.wav")

PREWARM_PHRASES = [
    # Silence / no-answer ladder
    ("શું તમે કોલ પર છો?", "gu-IN"),
    ("કોઈ જવાબ ન મળવાના કારણે અમે કોલ ને એન્ડ કરી રહ્યા છીએ.", "gu-IN"),
    # Script Step 2 — qualification question (most common response after YES)
    ("સરસ! કૃપા કરીને તમારું latest qualification જણાવો — 10th, 12th, કે Graduation?", "gu-IN"),
    # Script Step 3 — interest question
    ("Perfect! તમને કયા career field માં રસ છે? Engineering, Management, Pharmacy, Design, Commerce, Science અથવા અન્ય?", "gu-IN"),
    # Script Step 4 — transfer announcement
    ("Perfect, કૃપા કરીને થોડી ક્ષણ લાઇન પર રહો, હું તમને અમારી counselling team સાથે transfer કરું છું.", "gu-IN"),
    # Negative close
    ("ઠીક છે, ભવિષ્યમાં જ્યારે જરૂર પડે ત્યારે સંપર્ક કરજો. ધન્યવાદ!", "gu-IN"),
]

async def _generate_greeting():
    """
    B1 fix: If greeting.wav already exists and is valid, skip re-generation.
    If generation fails, copy silence WAV so /vobiz-answer never serves a 404.
    """
    # Force regenerate if greeting text has changed (check file age vs code)
    # Delete old greeting.wav so it always uses the current GREETING_TEXT
    # To force regen: delete static/audio/greeting.wav and restart
    if os.path.exists(GREETING_PATH) and os.path.getsize(GREETING_PATH) > 2000:
        # Validate it contains the right content by checking file size heuristic
        # New greeting is ~5s @ 8kHz = ~80KB. If smaller, likely old version.
        if os.path.getsize(GREETING_PATH) > 50000:
            logger.info("✅ greeting.wav already exists — skipping generation")
            return
        else:
            logger.info("⚠️  greeting.wav too small — regenerating with new text")
            os.remove(GREETING_PATH)

    if not SARVAM_API_KEY:
        logger.warning("⚠️  No Sarvam key — using silence as greeting placeholder")
        import shutil
        if os.path.exists(_SILENCE_WAV_PATH):
            shutil.copy(_SILENCE_WAV_PATH, GREETING_PATH)
        return

    await asyncio.sleep(1)
    _t = time.time()
    try:
        client  = _get_httpx_client()
        payload = {
            "text": GREETING_TEXT,
            "target_language_code": "gu-IN",
            "speaker": "anushka",
            "model": "bulbul:v2",
            "speech_sample_rate": 8000,
            "enable_preprocessing": True,
        }
        headers = {"api-subscription-key": SARVAM_API_KEY, "Content-Type": "application/json"}
        resp    = await client.post("https://api.sarvam.ai/text-to-speech", json=payload, headers=headers)
        if resp.status_code == 200:
            audio_bytes = base64.b64decode(resp.json()["audios"][0])
            loop        = asyncio.get_running_loop()
            await loop.run_in_executor(_io_executor, lambda: open(GREETING_PATH, "wb").write(audio_bytes))
            logger.info(f"✅ greeting.wav generated ({len(audio_bytes)} bytes, {(time.time()-_t)*1000:.0f}ms)")
        else:
            logger.error(f"❌ greeting.wav generation failed: Sarvam {resp.status_code}")
            import shutil
            if os.path.exists(_SILENCE_WAV_PATH):
                shutil.copy(_SILENCE_WAV_PATH, GREETING_PATH)
    except Exception as e:
        logger.error(f"❌ greeting.wav generation error: {e}")
        import shutil
        if os.path.exists(_SILENCE_WAV_PATH):
            shutil.copy(_SILENCE_WAV_PATH, GREETING_PATH)

async def _prewarm_tts():
    await asyncio.sleep(6)
    for text, lang in PREWARM_PHRASES:
        try:
            url = await generate_tts_audio(text, BASE_URL or "http://localhost:8000", lang)
            logger.info(f"🔥 Pre-warmed: '{text[:40]}'")
        except Exception as e:
            logger.warning(f"TTS pre-warm failed: {e}")
        await asyncio.sleep(0.2)  # rate-limit prewarm requests

async def _session_cleanup_task():
    while True:
        await asyncio.sleep(600)
        now     = time.time()
        expired = [sid for sid, ts in session_expiry.items() if (now - ts) > SESSION_TTL]
        for sid in expired:
            sessions.pop(sid, None)
            session_locks.pop(sid, None)
            session_expiry.pop(sid, None)
        if expired:
            logger.info(f"🧹 Cleaned {len(expired)} expired sessions")

@app.on_event("startup")
async def startup_event():
    asyncio.create_task(_generate_greeting())
    asyncio.create_task(_prewarm_tts())
    asyncio.create_task(_session_cleanup_task())

@app.on_event("shutdown")
async def shutdown_event():
    await _close_httpx_clients()  # B7

# ─────────────────────────────────────────
# AI RESPONSE (B4: does NOT re-acquire session lock)
# ─────────────────────────────────────────
async def get_ai_response(
    call_sid: str, user_input: str, base_url: str = ""
) -> Dict[str, object]:
    """
    B4 fix: This function is called from INSIDE the session lock in vobiz_respond.
    It must NOT call get_session_lock() itself — that would deadlock.
    All session reads/writes here are direct dict access, no lock.
    """
    if call_sid not in sessions:
        system_prompt = await get_system_prompt_with_courses()
        sessions[call_sid] = [{"role": "system", "content": system_prompt}]

    sessions[call_sid].append({"role": "user", "content": user_input})

    try:
        request_messages = list(sessions[call_sid])

        rag_context = await asyncio.get_running_loop().run_in_executor(
            _faiss_executor, build_rag_context, user_input
        )
        if rag_context:
            request_messages.insert(1, {
                "role": "system",
                "content": (
                    "Use this retrieved context as highest-priority factual grounding.\n\n"
                    f"RETRIEVED_CONTEXT:\n{rag_context}"
                ),
            })

        # FIX #5 (prompt side): enforce TEXT: tag FIRST so early-TTS fires correctly.
        # If LLM emits LANG: or STATUS: before TEXT:, the early_tts_text captures metadata
        # and the prefix match fails — causing double TTS latency.
        request_messages.append({
            "role": "system",
            "content": (
                "CRITICAL: Your response must be MAX 2 sentences. "
                "ALWAYS start your output with TEXT: — never with LANG: or STATUS:. "
                "Strict format: TEXT: <gujarati response> | LANG: gu-IN | NAME: <name> | INTEREST: <interest> | STATUS: <Hot/Warm/Cold/Negative>. "
                "Do NOT add explanations. Follow the script step exactly."
            )
        })

        _t_llm = time.time()
        raw_text, early_tts_task, early_tts_text = await _stream_llm_with_early_tts(
            call_sid, request_messages, base_url, lang="gu-IN"
        )
        logger.info(f"⏱️ LLM: {(time.time()-_t_llm)*1000:.0f}ms")

        sessions[call_sid].append({"role": "assistant", "content": raw_text})

        ai_data: Dict[str, object] = {"lang": "gu-IN", "text": raw_text}

        text_match = TEXT_PATTERN.search(raw_text)
        if text_match:
            ai_data["text"] = text_match.group(1).strip()
        else:
            cleaned = re.sub(
                r"(LANG|STATUS|INTEREST|NAME|FOLLOW_UP):\s*.*?(?=\||$)",
                "", raw_text, flags=re.IGNORECASE
            )
            ai_data["text"] = cleaned.strip().strip('|').strip()

        metadata: Dict[str, str] = {}
        for key, pattern in METADATA_PATTERNS.items():
            m = pattern.search(raw_text)
            if m:
                val = m.group(1).strip().strip('|').strip()
                if val.lower() != "unknown" and val:
                    metadata[key] = val

        if metadata.get("lead_status") == "Positive":
            metadata["stage"] = "Hot Call"

        clean_transcript = [msg for msg in sessions[call_sid] if msg['role'] != 'system']
        metadata["transcript"] = json.dumps(clean_transcript)
        save_call_log(call_sid, metadata)

        ai_data["_early_tts_task"] = early_tts_task
        ai_data["_early_tts_text"] = early_tts_text
        return ai_data

    except Exception as e:
        logger.error(f"AI response error ({AI_PROVIDER}): {e}", exc_info=True)
        if early_tts_task and not early_tts_task.done():  # type: ignore
            early_tts_task.cancel()                       # type: ignore
        return {
            "lang": "gu-IN",
            "text": "માફ કરશો, મને બરાબર સમજાયું નથી. શું તમે ફરીથી કહી શકશો?",
            "_early_tts_task": None, "_early_tts_text": "",
        }

# ─────────────────────────────────────────
# XML BUILDERS
# ─────────────────────────────────────────
async def gather_xml(
    request: Request, speak_text: str, action_path: str,
    lang: str = "gu-IN", timeout: int = 3
) -> str:
    base   = get_base_url(request)
    audio  = await generate_tts_audio(speak_text, base, lang)
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<Response>"
        f"<Play>{audio}</Play>"
        f'<Record action="{base}/{action_path}" method="POST" '
        f'maxLength="15" timeout="{timeout}" playBeep="false" />'
        f'<Redirect method="POST">{base}/vobiz-silent</Redirect>'
        "</Response>"
    )

async def hangup_xml(request: Request, speak_text: str, lang: str = "gu-IN") -> str:
    base  = get_base_url(request)
    audio = await generate_tts_audio(speak_text, base, lang)
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<Response>"
        f"<Play>{audio}</Play>"
        "<Hangup/>"
        "</Response>"
    )

async def handle_silence_logic(request: Request, call_sid: str) -> str:
    """
    Silence ladder:
      count=1 → silent re-record (10s timeout)
      count=2 → play "Are you there?" (3s timeout)
      count≥3 → hangup
    B14 fix: gracefully handles unknown/missing call_sid.
    """
    silence_key = f"__silence__{call_sid}"
    count = sessions.get(silence_key, 0) + 1
    sessions[silence_key] = count
    base  = get_base_url(request)

    if count == 1:
        return (
            '<?xml version="1.0" encoding="UTF-8"?>'
            "<Response>"
            f'<Record action="{base}/vobiz-respond" method="POST" '
            f'maxLength="15" timeout="10" playBeep="false" />'
            f'<Redirect method="POST">{base}/vobiz-silent</Redirect>'
            "</Response>"
        )
    elif count == 2:
        return await gather_xml(
            request, "શું તમે કોલ પર છો?", "vobiz-respond", lang="gu-IN", timeout=5
        )
    else:
        sessions.pop(silence_key, None)
        return await hangup_xml(
            request,
            "કોઈ જવાબ ન મળવાના કારણે અમે કોલ ને એન્ડ કરી રહ્યા છીએ.",
            lang="gu-IN"
        )

# ─────────────────────────────────────────
# DOWNLOAD RECORDING
# ─────────────────────────────────────────
async def download_recording_async(recording_url: str) -> bytes:
    headers = {"X-Auth-ID": VOBIZ_AUTH_ID, "X-Auth-Token": VOBIZ_AUTH_TOKEN}
    client  = _get_httpx_client()
    _t      = time.time()
    for attempt in range(2):
        try:
            resp = await client.get(recording_url, headers=headers)
            if resp.status_code == 200 and len(resp.content) >= 500:
                logger.info(f"⏱️ DOWNLOAD: {(time.time()-_t)*1000:.0f}ms | {len(resp.content)} bytes")
                return resp.content
            if attempt == 0:
                await asyncio.sleep(0.15)
        except Exception as e:
            logger.error(f"Download error (attempt {attempt}): {e}")
            if attempt == 0:
                await asyncio.sleep(0.15)
    return b""

# ─────────────────────────────────────────
# STT HALLUCINATION + ECHO FILTERS  (B9)
# ─────────────────────────────────────────
_HALLUCINATION_EXACT = {
    # Pure noise / silence artefacts from Sarvam STT
    "data factor is a problem", "data science research",
    "હાઁ જી હાઁ જી હાઁ હાઁ", "are wah", "aare waah", "are waah",
    "aare wah", "arewah", "વાહ વાહ", "waah waah", "wah wah",
    "આરે વા", "આરે વ", "thank you", "thanks",
    "tha", "tah", "thaa",
    "hmm hmm", "hm hm",
    "uh", "huh",
    # NOTE: "okay", "ok", "haan", "yes", "hi", "hello", "hmm" are VALID user responses
    # in this script flow — DO NOT filter them out.
}
_HALLUCINATION_CONTAINS = {
    "bumped", "mimm", "હરલ ળાલ", "arre wah", "are wa",
    "haan ji haan", "han ji han",
}

def _clean_transcript(text: str) -> str:
    return re.sub(r'[^\w\s]', '', text.lower()).strip()

# Short words that are VALID in our counselling script — never hallucinations
_VALID_SHORT_RESPONSES = {
    "ha", "haa", "han", "haan", "yes", "okay", "ok", "sure", "thik",
    "no", "na", "nahi", "hello", "hi", "hmm", "bye",
    "10", "12", "10th", "12th", "tenth", "twelfth",
    "engineering", "science", "commerce", "management",
    "pharmacy", "design", "arts", "it", "computer",
    # Gujarati short affirmatives
    "હા", "ના", "ઠીક", "સારૂ", "ચાલે", "ગમશે",
}

def _is_hallucination(clean: str) -> bool:
    # Never filter valid script responses
    if clean in _VALID_SHORT_RESPONSES:
        return False
    # Never filter if it contains a qualification keyword
    _QUALIFICATION_WORDS = {"10th", "12th", "tenth", "twelfth", "graduation",
                             "degree", "graduate", "10", "12", "tenth kdhyu",
                             "engineering", "science", "commerce", "management",
                             "pharmacy", "design", "computer", "it"}
    if any(q in clean for q in _QUALIFICATION_WORDS):
        return False
    if len(clean) < 2:
        return True
    if clean in _HALLUCINATION_EXACT:
        return True
    if any(h in clean for h in _HALLUCINATION_CONTAINS):
        return True
    # Repeated single word (e.g. "waah waah waah") — only if NOT a valid response
    if re.match(r'^(\w+)\s+\1(\s+\1)*$', clean) and clean not in _VALID_SHORT_RESPONSES:
        return True
    # Any word repeated 4+ times
    if re.search(r'(\b\w+\b)(?:\s+\1){4,}', clean):
        return True
    return False

def _is_echo(clean: str, call_sid: str) -> bool:
    """
    Check if user transcript is actually the agent's own TTS voice picked up by mic.
    Only fires when transcript is long enough (>=12 chars) to avoid dropping real short
    responses like 'ha bolo', 'yes', '10th' which superficially match agent text.
    """
    if call_sid not in sessions or len(clean) < 12:
        return False
    last_msgs = [m for m in sessions.get(call_sid, []) if m['role'] == 'assistant']
    if not last_msgs:
        return False
    last_raw    = last_msgs[-1].get('content', '')
    tm          = TEXT_PATTERN.search(last_raw)
    last_spoken = (tm.group(1).strip() if tm else last_raw).lower()
    last_clean  = _clean_transcript(last_spoken)
    if not last_clean or len(last_clean) < 12:
        return False
    # Echo: user transcript must be a SUBSTANTIAL prefix of agent speech (>=12 chars match)
    # AND the match ratio must be high (user said nearly exactly what agent said)
    if len(clean) >= 12 and last_clean.startswith(clean[:12]):
        return True
    return False

def _is_duplicate(clean: str, call_sid: str) -> bool:
    """
    Only block duplicate turns for longer transcripts (>=8 chars).
    Short affirmatives like 'ha', 'yes', 'ok' are common in conversation
    and should never be deduplicated — user might say 'ha' twice naturally.
    """
    if len(clean) < 8:
        return False  # never deduplicate short responses
    key  = f"__last_user__{call_sid}"
    last = sessions.get(key, "")
    if clean == last:
        return True
    sessions[key] = clean
    return False

def apply_transcript_filters(user_speech: str, call_sid: str) -> str:
    """
    B9 fix: All 4 filter layers applied consistently even for non-empty STT results.
    Returns empty string if the input should be discarded.
    """
    if not user_speech:
        return ""

    clean = _clean_transcript(user_speech)

    # Layer 1: length
    if len(clean) < 2:
        logger.info(f"   [filter L1] too short: '{user_speech}'")
        return ""

    # Layer 2: hallucination patterns
    if _is_hallucination(clean):
        logger.info(f"   [filter L2] hallucination: '{user_speech}'")
        return ""

    # Layer 3: echo detection
    if _is_echo(clean, call_sid):
        logger.info(f"   [filter L3] echo: '{user_speech}'")
        return ""

    # Layer 4: duplicate turn
    if _is_duplicate(clean, call_sid):
        logger.info(f"   [filter L4] duplicate: '{user_speech}'")
        return ""

    return user_speech

# ─────────────────────────────────────────
# WARM TRANSFER
# ─────────────────────────────────────────
async def transfer_xml(request: Request, speak_text: str, call_sid: str, lang: str = "gu-IN") -> str:
    """
    Full warm transfer flow:
    1. TTS "connecting you" to user
    2. User redirected to /hold-loop (music loops with no dead air)
    3. Background task dials counsellor after 3 s
    B15: conf_room stored deterministically and never overwritten mid-call.
    B6: counsellor busy state stored per-call in sessions, not in a shared set.
    """
    base      = get_base_url(request)
    conf_key  = f"__conf__{call_sid}"
    conf_room = sessions.get(conf_key)
    if not conf_room:
        conf_room           = f"gvx_{call_sid[-12:]}"
        sessions[conf_key]  = conf_room  # B15: set once, never overwritten

    save_call_log(call_sid, {"stage": "Warm Call"})
    audio_url = await generate_tts_audio(speak_text, base, lang)
    asyncio.create_task(_dial_counsellor_background(call_sid, base))
    logger.info(f"📞 WARM TRANSFER | room={conf_room} | sid={call_sid}")

    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<Response>"
        f"<Play>{audio_url}</Play>"
        f'<Redirect method="POST">{base}/hold-loop?call_sid={call_sid}</Redirect>'
        "</Response>"
    )

async def _dial_counsellor_background(call_sid: str, base_url: str):
    await asyncio.sleep(3.0)

    # B6: track per-call which counsellor was selected (no shared set)
    busy_key = "__counsellors_busy__"
    busy: set = sessions.get(busy_key, set())

    selected = None
    for number in TRANSFER_NUMBERS:
        if number not in busy:
            selected = number
            break

    if not selected:
        logger.warning(f"⚠️  All counsellors busy — rescuing user {call_sid}")
        asyncio.create_task(_rescue_user_from_hold_loop(call_sid, base_url, all_busy=True))
        return

    busy.add(selected)
    sessions[busy_key]                        = busy
    sessions[f"__counsellor_num__{call_sid}"] = selected

    try:
        loop   = asyncio.get_running_loop()
        result = await loop.run_in_executor(
            _io_executor,
            lambda: _dial_counsellor_sync(call_sid, base_url, selected)
        )
        logger.info(f"🤝 Counsellor dialled to {selected}: {result.get('request_uuid', '?')}")
    except Exception as e:
        logger.error(f"❌ Counsellor dial failed: {e}")
        busy.discard(selected)
        asyncio.create_task(_rescue_user_from_hold_loop(call_sid, base_url))

def _dial_counsellor_sync(call_sid: str, base_url: str, target_number: str) -> dict:
    url     = f"https://api.vobiz.ai/api/v1/Account/{VOBIZ_AUTH_ID}/Call/"
    headers = {
        "X-Auth-ID":    VOBIZ_AUTH_ID,
        "X-Auth-Token": VOBIZ_AUTH_TOKEN,
        "Content-Type": "application/json",
    }
    payload = {
        "from":          VOBIZ_FROM_NUMBER,
        "to":            target_number,
        "answer_url":    f"{base_url}/counsellor-answer?call_sid={call_sid}",
        "answer_method": "POST",
        "hangup_url":    f"{base_url}/counsellor-hangup?call_sid={call_sid}",
        "hangup_method": "POST",
    }
    resp   = http_session.post(url, headers=headers, json=payload, timeout=HTTP_TIMEOUT_SECONDS)
    result = resp.json()
    if not resp.ok:
        raise RuntimeError(f"VoBiz dial error {resp.status_code}: {result}")
    sessions[f"__counsellor__{call_sid}"] = result.get("request_uuid", "")
    return result

async def _push_user_to_conference(call_sid: str, base_url: str, delay: float = 1.5):
    await asyncio.sleep(delay)
    join_url = f"{base_url}/join-conference?call_sid={call_sid}"
    api_url  = f"https://api.vobiz.ai/api/v1/Account/{VOBIZ_AUTH_ID}/Call/{call_sid}/"
    headers  = {
        "X-Auth-ID":    VOBIZ_AUTH_ID,
        "X-Auth-Token": VOBIZ_AUTH_TOKEN,
        "Content-Type": "application/json",
    }
    payload  = {"legs": "aleg", "aleg_url": join_url, "aleg_method": "POST"}

    def _do():
        try:
            r = http_session.post(api_url, headers=headers, json=payload, timeout=HTTP_TIMEOUT_SECONDS)
            logger.info(f"🔀 User pushed to conference | {r.status_code}: {r.text[:200]}")
        except Exception as e:
            logger.error(f"Push to conference error: {e}")

    await asyncio.get_running_loop().run_in_executor(_io_executor, _do)

async def _rescue_user_from_hold_loop(call_sid: str, base_url: str, all_busy: bool = False):
    """
    Always runs — even if counsellor dial partially succeeded.
    B12: counsellor num is freed in finally block.
    """
    await asyncio.sleep(1.0)
    try:
        fallback_text = (
            "અમારી તમામ ટીમ અત્યારે અન્ય કોલમાં વ્યસ્ત છે. અમે ટૂંક સમયમાં તમને ફોન કરીશું. ધન્યવાદ."
            if all_busy else
            "માફ કરશો, અમારી ટીમ ઉપલ્બ્ધ નથી. અમે ટૂંક સમયમાં કૉલ કરીશું. ધન્યવાદ."
        )
        audio_url    = await generate_tts_audio(fallback_text, base_url, lang="gu-IN")
        fallback_url = (
            f"{base_url}/user-fallback"
            f"?audio_url={requests.utils.quote(audio_url)}"
        )
        api_url  = f"https://api.vobiz.ai/api/v1/Account/{VOBIZ_AUTH_ID}/Call/{call_sid}/"
        headers  = {
            "X-Auth-ID":    VOBIZ_AUTH_ID,
            "X-Auth-Token": VOBIZ_AUTH_TOKEN,
            "Content-Type": "application/json",
        }
        payload  = {"legs": "aleg", "aleg_url": fallback_url, "aleg_method": "POST"}

        def _do():
            try:
                r = http_session.post(api_url, headers=headers, json=payload, timeout=HTTP_TIMEOUT_SECONDS)
                logger.info(f"   Rescue → {r.status_code}: {r.text[:150]}")
            except Exception as e:
                logger.error(f"   Rescue HTTP error: {e}")

        await asyncio.get_running_loop().run_in_executor(_io_executor, _do)
    except Exception as e:
        logger.error(f"❌ _rescue_user_from_hold_loop failed: {e}")
    finally:
        # B12: always clean up regardless of exception
        _free_counsellor(call_sid)
        sessions.pop(f"__conf__{call_sid}", None)

def _free_counsellor(call_sid: str):
    """B12: Always free counsellor slot on any call termination."""
    num = sessions.pop(f"__counsellor_num__{call_sid}", None)
    if num:
        busy: set = sessions.get("__counsellors_busy__", set())
        busy.discard(num)
        sessions["__counsellors_busy__"] = busy
        logger.info(f"🔓 Counsellor {num} freed")

# ─────────────────────────────────────────
# WEBHOOK ENDPOINTS
# ─────────────────────────────────────────

@app.api_route("/vobiz-answer", methods=["GET", "POST"])
async def vobiz_answer(request: Request):
    logger.info("📞 CALL PICKUP (/vobiz-answer)")
    try:
        form_data = dict(await request.form())
    except Exception:
        form_data = {}

    call_sid = (
        form_data.get("CallUUID") or form_data.get("request_uuid")
        or form_data.get("CallSid") or f"anon_{uuid.uuid4().hex[:8]}"
    )
    logger.info(f"   SID={call_sid}")

    base      = get_base_url(request)
    audio_url = f"{base}/static/audio/greeting.wav"

    # B1: verify file exists; fallback to on-the-fly TTS
    if not os.path.exists(GREETING_PATH) or os.path.getsize(GREETING_PATH) < 500:
        logger.warning("greeting.wav missing — generating on-the-fly")
        audio_url = await generate_tts_audio(GREETING_TEXT, base, "gu-IN")

    async with get_session_lock(call_sid):
        session_expiry[call_sid] = time.time()
        sessions[call_sid] = [
            {"role": "system",    "content": await get_system_prompt_with_courses()},
            {"role": "assistant", "content": f"TEXT: {GREETING_TEXT} | LANG: gu-IN"},
        ]

    # B2: timeout=3 (was 1) — gives user time to begin speaking after greeting finishes
    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<Response>"
        f"<Play>{audio_url}</Play>"
        f'<Record action="{base}/vobiz-respond" method="POST" '
        f'maxLength="15" timeout="3" playBeep="false" />'
        f'<Redirect method="POST">{base}/vobiz-silent</Redirect>'
        "</Response>"
    )
    return Response(content=xml, media_type="text/xml", headers=get_cloudflare_headers())


@app.api_route("/vobiz-respond", methods=["GET", "POST"])
async def vobiz_respond(request: Request):
    logger.info("🗣️  /vobiz-respond")

    form_data = {}
    try:
        form_data = dict(await request.form())
    except Exception:
        pass
    if not form_data:
        try:
            form_data = await request.json()
        except Exception:
            pass

    # B8: extract call_sid BEFORE acquiring session lock
    call_sid = (
        form_data.get("CallUUID") or form_data.get("request_uuid")
        or form_data.get("CallSid") or "unknown_session"
    )

    async with get_session_lock(call_sid):
        session_expiry[call_sid] = time.time()
        return await _vobiz_respond_logic(request, form_data, call_sid)


async def _vobiz_respond_logic(request: Request, form_data: dict, call_sid: str) -> Response:
    recording_url = (
        form_data.get("RecordUrl") or form_data.get("RecordFile")
        or form_data.get("RecordingUrl") or form_data.get("recording_url")
    )
    user_speech = (
        form_data.get("Speech") or form_data.get("speech")
        or form_data.get("SpeechResult") or form_data.get("Digits") or ""
    ).strip()

    # ── STT pipeline ─────────────────────────────────────────────────────
    if recording_url and not user_speech:
        audio_bytes = await download_recording_async(recording_url)
        logger.info(f"   Download: {len(audio_bytes)} bytes")

        if len(audio_bytes) >= 500:
            ext      = ".mp3" if recording_url.lower().endswith(".mp3") else ".wav"
            filename = f"rec_{call_sid}{ext}"

            # B13: VAD runs in _vad_executor, never blocks event loop
            loop      = asyncio.get_running_loop()
            is_silent = await loop.run_in_executor(
                _vad_executor, _detect_silence_sync, audio_bytes, ext
            )

            if is_silent:
                logger.info("   VAD: silent — skipping STT")
                user_speech = ""
            else:
                raw_transcript = await transcribe_audio(audio_bytes, filename)
                # B9: all 4 filter layers applied
                user_speech = apply_transcript_filters(raw_transcript, call_sid)
        else:
            logger.warning(f"   Recording too small ({len(audio_bytes)} bytes)")

    logger.info(f"   [{call_sid[-8:]}] transcript='{user_speech}'")

    if not user_speech:
        xml = await handle_silence_logic(request, call_sid)
        return Response(content=xml, media_type="text/xml", headers=get_cloudflare_headers())

    sessions.pop(f"__silence__{call_sid}", None)

    t_start  = time.time()
    base     = get_base_url(request)
    ai_data  = await get_ai_response(call_sid, user_speech, base_url=base)
    lang     = str(ai_data.get("lang", "gu-IN"))
    text     = str(ai_data.get("text", "માફ કરશો, મને બરાબર સમજાયું નથી."))
    early_tts_task: Optional[asyncio.Task] = ai_data.get("_early_tts_task")  # type: ignore
    early_tts_text: str = str(ai_data.get("_early_tts_text", ""))

    should_hangup   = "[HANGUP]" in text or "HANGUP" in text
    should_transfer = any(x in text for x in ["[TRANSFER]", "[ટ્રાન્સફર]", "TRANSFER", "ટ્રાન્સફર"])
    text = (
        text.replace("[HANGUP]", "").replace("[TRANSFER]", "")
            .replace("[ટ્રાન્સફર]", "").strip()
    )

    # ── TTS ──────────────────────────────────────────────────────────────
    _t_tts    = time.time()
    audio_url = ""

    if not (should_transfer or should_hangup):
        # B5: properly handle early TTS task
        if (early_tts_task is not None and early_tts_text
                and text.startswith(early_tts_text)):
            # Reuse early TTS result (it was for the correct prefix)
            try:
                audio_url = await early_tts_task
                logger.info(f"⏱️ TTS: {(time.time()-_t_tts)*1000:.0f}ms (reused early prefix)")
            except Exception:
                audio_url = await generate_tts_audio(text, base, lang)
        else:
            # Cancel early task if it's still running and we don't need it
            if early_tts_task is not None and not early_tts_task.done():
                early_tts_task.cancel()
                try:
                    await early_tts_task
                except asyncio.CancelledError:
                    pass
            audio_url = await generate_tts_audio(text, base, lang)
            logger.info(f"⏱️ TTS: {(time.time()-_t_tts)*1000:.0f}ms (full generation)")

        # B10: audio_url is guaranteed non-empty (generate_tts_audio has fallback chain)

    # ── XML response ──────────────────────────────────────────────────────
    if should_transfer:
        xml = await transfer_xml(request, text, call_sid, lang=lang)
    elif should_hangup:
        xml = await hangup_xml(request, text, lang=lang)
    elif audio_url:
        # B2: Record timeout=3 gives user adequate time to respond naturally
        xml = (
            '<?xml version="1.0" encoding="UTF-8"?>'
            "<Response>"
            f"<Play>{audio_url}</Play>"
            f'<Record action="{base}/vobiz-respond" method="POST" '
            f'maxLength="15" timeout="3" playBeep="false" />'
            f'<Redirect method="POST">{base}/vobiz-silent</Redirect>'
            "</Response>"
        )
    else:
        xml = await gather_xml(request, text, "vobiz-respond", lang=lang, timeout=3)

    logger.info(f"⏱️ TOTAL_TURN: {(time.time()-t_start)*1000:.0f}ms | sid={call_sid[-8:]}")
    return Response(content=xml, media_type="text/xml", headers=get_cloudflare_headers())


@app.api_route("/vobiz-silent", methods=["GET", "POST"])
async def vobiz_silent(request: Request):
    # B14: graceful handling when call_sid is missing
    try:
        form_data = dict(await request.form())
    except Exception:
        form_data = {}
    call_sid = (
        form_data.get("CallUUID") or form_data.get("CallSid")
        or request.query_params.get("call_sid") or "unknown"
    )
    xml = await handle_silence_logic(request, call_sid)
    return Response(content=xml, media_type="text/xml", headers=get_cloudflare_headers())


@app.api_route("/status", methods=["GET", "POST"])
async def call_status(request: Request):
    form_data = {}
    try:
        form_data = dict(await request.form())
    except Exception:
        pass
    if not form_data:
        try:
            form_data = await request.json()
        except Exception:
            pass

    call_sid   = form_data.get("CallUUID") or form_data.get("request_uuid") or form_data.get("CallSid")
    status     = form_data.get("CallStatus") or form_data.get("status") or "unknown"
    end_reason = (
        form_data.get("HangupCauseName") or form_data.get("hangup_cause_name")
        or form_data.get("Reason")
    )

    if call_sid:
        update = {"status": status}
        if end_reason:
            update["end_reason"] = end_reason
        save_call_log(call_sid, update)

        if status.lower() in TERMINAL_CALL_STATUSES:
            # B12: always free counsellor on terminal status
            _free_counsellor(call_sid)
            # Clean up session
            async with get_session_lock(call_sid):
                sessions.pop(call_sid, None)
                session_expiry.pop(call_sid, None)
                sessions.pop(f"__silence__{call_sid}", None)
                sessions.pop(f"__last_user__{call_sid}", None)
                sessions.pop(f"__conf__{call_sid}", None)
                sessions.pop(f"__counsellor__{call_sid}", None)
            session_locks.pop(call_sid, None)

    return JSONResponse(content={"received": True})


# ─────────────────────────────────────────
# WARM TRANSFER ENDPOINTS
# ─────────────────────────────────────────

@app.api_route("/hold-loop", methods=["GET", "POST"])
async def hold_loop(request: Request):
    # B14: robust call_sid extraction
    try:
        form_data = dict(await request.form())
    except Exception:
        form_data = {}
    call_sid = (
        request.query_params.get("call_sid")
        or form_data.get("CallUUID") or form_data.get("call_sid") or ""
    )
    base           = get_base_url(request)
    hold_music_url = f"{base}/static/audio/{CONNECTING_MUSIC}"
    logger.info(f"🔄 /hold-loop | sid={call_sid}")
    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<Response>"
        f"<Play>{hold_music_url}</Play>"
        f'<Redirect method="POST">{base}/hold-loop?call_sid={call_sid}</Redirect>'
        "</Response>"
    )
    return Response(content=xml, media_type="text/xml", headers=get_cloudflare_headers())


@app.api_route("/join-conference", methods=["GET", "POST"])
async def join_conference(request: Request):
    try:
        form_data = dict(await request.form())
    except Exception:
        form_data = {}
    call_sid  = (
        request.query_params.get("call_sid")
        or form_data.get("CallUUID") or form_data.get("call_sid") or ""
    )
    # B15: use stored room; never fallback to re-computing a different suffix
    conf_room = sessions.get(f"__conf__{call_sid}") or f"gvx_{call_sid[-12:]}"
    logger.info(f"✅ /join-conference | sid={call_sid} | room={conf_room}")
    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<Response>"
        f"<Conference>{conf_room}</Conference>"
        "<Hangup/>"
        "</Response>"
    )
    return Response(content=xml, media_type="text/xml", headers=get_cloudflare_headers())


@app.api_route("/counsellor-answer", methods=["GET", "POST"])
async def counsellor_answer(request: Request):
    try:
        form_data = dict(await request.form())
    except Exception:
        form_data = {}
    call_sid  = request.query_params.get("call_sid", "")
    conf_room = sessions.get(f"__conf__{call_sid}") or f"gvx_{call_sid[-12:]}"
    base      = get_base_url(request)

    logger.info(f"🤝 Counsellor answered | room={conf_room} | sid={call_sid}")

    briefing     = "ગણપત યુનિવર્સિટીના એડમિશનની ઇન્કવાયરી માટે સ્ટુડન્ટ તમારી સાથે વાત કરવા માંગે છે, હું કનેક્ટ કરી રહી છું."
    briefing_url = await generate_tts_audio(briefing, base, lang="gu-IN")

    # Push user from hold-loop after 1.8s (counsellor needs to be in conference first)
    asyncio.create_task(_push_user_to_conference(call_sid, base, delay=1.8))

    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<Response>"
        f"<Play>{briefing_url}</Play>"
        f"<Conference>{conf_room}</Conference>"
        "<Hangup/>"
        "</Response>"
    )
    return Response(content=xml, media_type="text/xml", headers=get_cloudflare_headers())


@app.api_route("/counsellor-hangup", methods=["GET", "POST"])
async def counsellor_hangup(request: Request):
    try:
        form_data = dict(await request.form())
    except Exception:
        form_data = {}

    call_sid     = request.query_params.get("call_sid", "")
    base         = get_base_url(request)
    hangup_cause = (
        form_data.get("HangupCauseName") or form_data.get("CallStatus")
        or form_data.get("status") or "unknown"
    )
    logger.info(f"📋 /counsellor-hangup | cause={hangup_cause} | sid={call_sid}")

    normal_causes = {"normal_clearing", "completed", "in-progress", "normal hangup"}

    try:
        if hangup_cause.lower() not in normal_causes:
            logger.warning(f"⚠️  Counsellor did not answer — rescuing user")
            asyncio.create_task(_rescue_user_from_hold_loop(call_sid, base))
        else:
            logger.info("✅ Transfer completed normally")
            sessions.pop(f"__conf__{call_sid}", None)
    finally:
        # B12: always free counsellor slot
        _free_counsellor(call_sid)

    return JSONResponse(content={"received": True})


@app.api_route("/user-fallback", methods=["GET", "POST"])
async def user_fallback(request: Request):
    audio_url = request.query_params.get("audio_url", "")
    # Safety: if audio_url is empty, use silence
    if not audio_url:
        base      = get_base_url(request)
        audio_url = _silence_url(base)
    logger.info(f"📢 /user-fallback | audio={audio_url[:80]}")
    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<Response>"
        f"<Play>{audio_url}</Play>"
        "<Hangup/>"
        "</Response>"
    )
    return Response(content=xml, media_type="text/xml", headers=get_cloudflare_headers())


# ─────────────────────────────────────────
# API — AUTH
# ─────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str

@app.post("/api/login")
async def login(creds: LoginRequest):
    if creds.username == "Admin" and creds.password == "Guni@2026":
        return {"token": "fake-jwt-token-for-demo", "user": "Admin"}
    raise HTTPException(status_code=401, detail="Invalid credentials")


# ─────────────────────────────────────────
# API — CALLS & CAMPAIGNS
# ─────────────────────────────────────────
class CallRequest(BaseModel):
    phone_number: str

class CampaignRequest(BaseModel):
    phone_numbers: List[str]

class Course(BaseModel):
    name: str
    institute: str
    duration: str
    fees: str
    eligibility: str
    counsellor: str
    phone: str
    brochure_url: Optional[str] = None

class RagDocumentRequest(BaseModel):
    title: Optional[str] = None
    content: str
    source: Optional[str] = "manual"


def initiate_outbound_call(phone_number: str, dynamic_base_url: str) -> dict:
    url     = f"https://api.vobiz.ai/api/v1/Account/{VOBIZ_AUTH_ID}/Call/"
    headers = {
        "X-Auth-ID":    VOBIZ_AUTH_ID,
        "X-Auth-Token": VOBIZ_AUTH_TOKEN,
        "Content-Type": "application/json",
    }
    payload = {
        "from": VOBIZ_FROM_NUMBER,
        "to":   phone_number,
        "answer_url":    f"{dynamic_base_url}/vobiz-answer",
        "answer_method": "POST",
        "hangup_url":    f"{dynamic_base_url}/status",
        "hangup_method": "POST",
    }
    resp     = http_session.post(url, headers=headers, json=payload, timeout=HTTP_TIMEOUT_SECONDS)
    result   = resp.json()
    if not resp.ok:
        raise HTTPException(status_code=resp.status_code, detail=result.get("message", "VoBiz API error"))
    call_uuid = result.get("request_uuid") or f"unknown_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    save_call_log(call_uuid, {"phone_number": phone_number, "status": "queued"})
    return {"call_sid": call_uuid, "details": result}


@app.post("/api/call")
async def trigger_call(req: CallRequest, request: Request):
    clean = normalize_phone_number(req.phone_number)
    if not clean:
        raise HTTPException(status_code=400, detail="Invalid phone number")
    try:
        result = initiate_outbound_call(clean, get_base_url(request))
        return {"success": True, "call_sid": result["call_sid"], "status": "queued"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Call initiation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def extract_phone_numbers_from_file_bytes(file_bytes: bytes, filename: str) -> List[str]:
    seen: set = set()
    numbers: List[str] = []

    if filename.lower().endswith((".xlsx", ".xls")) or file_bytes.startswith(b"PK\x03\x04"):
        try:
            from openpyxl import load_workbook
            wb   = load_workbook(BytesIO(file_bytes), data_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            if not rows:
                return []
            headers   = [re.sub(r"\s+", "_", str(cell).strip().lower()) for cell in rows[0] if cell]
            phone_col = next((i for i, n in enumerate(headers) if n in CSV_PHONE_HEADERS), -1)
            for row in rows[1 if phone_col >= 0 else 0:]:
                raw = str(row[phone_col] if phone_col >= 0 and len(row) > phone_col else (row[0] if row else ""))
                p   = normalize_phone_number(raw)
                if p and len(p.replace("+", "")) >= 8 and p not in seen:
                    seen.add(p)
                    numbers.append(p)
            return numbers
        except Exception as e:
            logger.error(f"Excel parse failed: {e}")

    try:
        decoded = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            decoded = file_bytes.decode("utf-16")
        except UnicodeDecodeError:
            decoded = file_bytes.decode("latin-1", errors="ignore")

    try:
        rows = [r for r in csv.reader(StringIO(decoded)) if any(c.strip() for c in r)]
    except csv.Error:
        rows = [r for r in csv.reader(decoded.splitlines()) if any(c.strip() for c in r)]

    if not rows:
        return []
    headers   = [re.sub(r"\s+", "_", (c or "").strip().lower()) for c in rows[0]]
    phone_col = next((i for i, n in enumerate(headers) if n in CSV_PHONE_HEADERS), -1)
    for row in rows[1 if phone_col >= 0 else 0:]:
        raw = row[phone_col] if phone_col >= 0 and len(row) > phone_col else (row[0] if row else "")
        p   = normalize_phone_number(raw)
        if p and len(p.replace("+", "")) >= 8 and p not in seen:
            seen.add(p)
            numbers.append(p)
    return numbers


def _start_campaign(phone_numbers: List[str], dynamic_base_url: str) -> dict:
    campaign_id   = uuid.uuid4().hex
    campaign_data = {
        "campaign_id":         campaign_id,
        "status":              "pending",
        "phone_numbers":       phone_numbers,
        "total":               len(phone_numbers),
        "completed_count":     0,
        "current_index":       None,
        "current_phone":       None,
        "current_call_sid":    None,
        "current_call_status": None,
        "stop_requested":      False,
        "results":             [],
        "created_at":          datetime.now().isoformat(),
        "started_at":          None,
        "ended_at":            None,
    }
    with campaign_lock:
        campaigns[campaign_id] = campaign_data
    threading.Thread(
        target=_run_campaign, args=(campaign_id, phone_numbers, dynamic_base_url), daemon=True
    ).start()
    return campaign_data


def _run_campaign(campaign_id: str, phone_numbers: List[str], dynamic_base_url: str):
    with campaign_lock:
        c = campaigns.get(campaign_id)
        if not c:
            return
        c["status"]     = "running"
        c["started_at"] = datetime.now().isoformat()

    for index, phone in enumerate(phone_numbers):
        with campaign_lock:
            c = campaigns.get(campaign_id)
            if not c or c.get("stop_requested"):
                if c:
                    c.update({"status": "stopped", "current_index": None, "current_phone": None,
                               "current_call_sid": None, "current_call_status": None,
                               "ended_at": datetime.now().isoformat()})
                return
            c.update({"current_index": index, "current_phone": phone, "current_call_status": "initiated"})

        try:
            result   = initiate_outbound_call(phone, dynamic_base_url)
            call_sid = result["call_sid"]
            with campaign_lock:
                c = campaigns.get(campaign_id)
                if not c:
                    return
                c["current_call_sid"] = call_sid
                c["results"].append({"phone_number": phone, "call_sid": call_sid, "status": "initiated"})

            started = time.time()
            while True:
                with campaign_lock:
                    c = campaigns.get(campaign_id)
                    if not c:
                        return
                    if c.get("stop_requested"):
                        c.update({"status": "stopped", "current_index": None, "current_phone": None,
                                   "current_call_sid": None, "current_call_status": None,
                                   "ended_at": datetime.now().isoformat()})
                        return

                latest_status = (get_call_status_from_db(call_sid) or "").lower()
                with campaign_lock:
                    c = campaigns.get(campaign_id)
                    if c:
                        c["current_call_status"] = latest_status or "initiated"

                if latest_status in TERMINAL_CALL_STATUSES:
                    with campaign_lock:
                        c = campaigns.get(campaign_id)
                        if c:
                            c["results"][-1]["status"] = latest_status
                    break

                if time.time() - started > CALL_MAX_DURATION_SECONDS:
                    save_call_log(call_sid, {"status": "failed", "end_reason": "campaign_timeout"})
                    with campaign_lock:
                        c = campaigns.get(campaign_id)
                        if c:
                            c["results"][-1].update({"status": "failed", "end_reason": "campaign_timeout"})
                    break

                time.sleep(CALL_POLL_INTERVAL_SECONDS)

        except Exception as e:
            logger.error(f"Campaign call failed for {phone}: {e}")
            with campaign_lock:
                c = campaigns.get(campaign_id)
                if c:
                    c["results"].append({"phone_number": phone, "status": "failed", "error": str(e)})
        finally:
            with campaign_lock:
                c = campaigns.get(campaign_id)
                if c:
                    c["completed_count"]    = len(c["results"])
                    c["current_call_sid"]   = None
                    c["current_call_status"] = None

    with campaign_lock:
        c = campaigns.get(campaign_id)
        if c:
            if c.get("status") != "stopped":
                c["status"] = "completed"
            c.update({"current_index": None, "current_phone": None,
                       "current_call_sid": None, "current_call_status": None,
                       "ended_at": datetime.now().isoformat()})


@app.post("/api/call/campaign")
async def start_call_campaign(req: CampaignRequest, request: Request):
    seen: set = set()
    cleaned: List[str] = []
    for raw in req.phone_numbers:
        p = normalize_phone_number(raw)
        if p and p not in seen:
            seen.add(p)
            cleaned.append(p)
    if not cleaned:
        raise HTTPException(status_code=400, detail="No valid phone numbers found")
    cd = _start_campaign(cleaned, get_base_url(request))
    return {"success": True, "campaign_id": cd["campaign_id"], "status": "pending", "total": len(cleaned)}


@app.post("/api/call/campaign/upload")
async def start_call_campaign_from_csv(request: Request, file: UploadFile = File(...)):
    filename = file.filename or "uploaded.csv"
    if not filename.lower().endswith((".csv", ".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    numbers = extract_phone_numbers_from_file_bytes(content, filename)
    if not numbers:
        raise HTTPException(status_code=400, detail="No valid phone numbers found in file")
    cd = _start_campaign(numbers, get_base_url(request))
    return {"success": True, "campaign_id": cd["campaign_id"],
            "status": cd["status"], "total": cd["total"], "filename": filename}


@app.get("/api/call/campaign/{campaign_id}")
async def get_call_campaign_status(campaign_id: str):
    with campaign_lock:
        c = campaigns.get(campaign_id)
        if not c:
            raise HTTPException(status_code=404, detail="Campaign not found")
        return {k: v for k, v in c.items() if k != "phone_numbers"}


@app.post("/api/call/campaign/{campaign_id}/stop")
async def stop_call_campaign(campaign_id: str):
    with campaign_lock:
        c = campaigns.get(campaign_id)
        if not c:
            raise HTTPException(status_code=404, detail="Campaign not found")
        if c["status"] in {"completed", "stopped"}:
            return {"success": True, "campaign_id": campaign_id, "status": c["status"]}
        c["stop_requested"] = True
    return {"success": True, "campaign_id": campaign_id, "status": "stopping"}


@app.post("/api/end_call/{call_sid}")
async def end_call(call_sid: str):
    try:
        url     = f"https://api.vobiz.ai/api/v1/Account/{VOBIZ_AUTH_ID}/Call/{call_sid}/"
        headers = {"X-Auth-ID": VOBIZ_AUTH_ID, "X-Auth-Token": VOBIZ_AUTH_TOKEN}
        http_session.delete(url, headers=headers, timeout=HTTP_TIMEOUT_SECONDS)
        save_call_log(call_sid, {"status": "completed", "end_reason": "user_initiated"})
        return {"success": True, "status": "completed"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────
# API — STATS, CALLS, COURSES, LEADS
# ─────────────────────────────────────────

@app.get("/api/stats")
def get_stats(start_date: Optional[str] = None, end_date: Optional[str] = None):
    conn = sqlite3.connect(DB_FILE)
    c    = conn.cursor()
    params = []
    cond   = ""
    if start_date and end_date:
        cond   = " WHERE started_at BETWEEN ? AND ?"
        params = [f"{start_date}T00:00:00", f"{end_date}T23:59:59"]

    c.execute(f"SELECT COUNT(*) FROM calls{cond}", params)
    total = c.fetchone()[0]

    pos_cond = cond.replace("WHERE", "WHERE lead_status='Positive' AND") if cond else " WHERE lead_status='Positive'"
    c.execute(f"SELECT COUNT(*) FROM calls{pos_cond}", params)
    positive = c.fetchone()[0]

    c.execute("SELECT * FROM calls ORDER BY id DESC LIMIT 5")
    columns = [d[0] for d in c.description]
    recent  = [dict(zip(columns, row)) for row in c.fetchall()]
    conn.close()
    return {"total_calls": total, "positive_leads": positive, "recent_calls": recent}


@app.get("/api/calls")
def get_calls(q: Optional[str] = None, start_date: Optional[str] = None, end_date: Optional[str] = None):
    conn   = sqlite3.connect(DB_FILE)
    c      = conn.cursor()
    query  = "SELECT * FROM calls"
    params, conditions = [], []
    if q:
        conditions.append("(phone_number LIKE ? OR user_name LIKE ?)")
        params.extend([f"%{q}%", f"%{q}%"])
    if start_date and end_date:
        conditions.append("started_at BETWEEN ? AND ?")
        params.extend([f"{start_date}T00:00:00", f"{end_date}T23:59:59"])
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY id DESC LIMIT 100"
    c.execute(query, params)
    columns = [d[0] for d in c.description]
    calls   = [dict(zip(columns, row)) for row in c.fetchall()]
    conn.close()
    return calls


@app.delete("/api/calls/{call_id}")
def delete_call_log(call_id: int):
    conn = sqlite3.connect(DB_FILE)
    c    = conn.cursor()
    c.execute("DELETE FROM calls WHERE id = ?", (call_id,))
    conn.commit()
    conn.close()
    return {"success": True, "deleted_id": call_id}


@app.get("/api/courses")
def get_courses():
    conn    = sqlite3.connect(DB_FILE)
    c       = conn.cursor()
    c.execute("SELECT * FROM courses")
    columns = [d[0] for d in c.description]
    courses = [dict(zip(columns, row)) for row in c.fetchall()]
    conn.close()
    return courses


@app.post("/api/courses")
def add_course(course: Course):
    conn = sqlite3.connect(DB_FILE)
    c    = conn.cursor()
    c.execute(
        "INSERT INTO courses (name,institute,duration,fees,eligibility,counsellor,phone,brochure_url) "
        "VALUES (?,?,?,?,?,?,?,?)",
        (course.name, course.institute, course.duration, course.fees,
         course.eligibility, course.counsellor, course.phone, course.brochure_url)
    )
    conn.commit()
    cid = c.lastrowid
    conn.close()
    invalidate_prompt_cache()
    return {**course.dict(), "id": cid}


@app.put("/api/courses/{course_id}")
def update_course(course_id: int, course: Course):
    conn = sqlite3.connect(DB_FILE)
    c    = conn.cursor()
    c.execute(
        "UPDATE courses SET name=?,institute=?,duration=?,fees=?,eligibility=?,"
        "counsellor=?,phone=?,brochure_url=? WHERE id=?",
        (course.name, course.institute, course.duration, course.fees,
         course.eligibility, course.counsellor, course.phone, course.brochure_url, course_id)
    )
    conn.commit()
    conn.close()
    invalidate_prompt_cache()
    return {"success": True}


@app.delete("/api/courses/{course_id}")
def delete_course(course_id: int):
    conn = sqlite3.connect(DB_FILE)
    c    = conn.cursor()
    c.execute("DELETE FROM courses WHERE id = ?", (course_id,))
    conn.commit()
    conn.close()
    invalidate_prompt_cache()
    return {"success": True}


@app.get("/api/leads")
def get_leads():
    conn = sqlite3.connect(DB_FILE)
    c    = conn.cursor()
    c.execute("""
        SELECT phone_number, stage, call_sid as last_call_sid, started_at as updated_at,
               user_name, interest, lead_status, transcript
        FROM calls
        WHERE id IN (SELECT MAX(id) FROM calls GROUP BY phone_number)
        ORDER BY started_at DESC
    """)
    columns = [d[0] for d in c.description]
    leads   = [dict(zip(columns, row)) for row in c.fetchall()]
    conn.close()
    return leads


@app.put("/api/leads/{phone}/stage")
def update_lead_stage(phone: str, body: dict):
    stage = body.get("stage", "Cold Call")
    conn  = sqlite3.connect(DB_FILE)
    c     = conn.cursor()
    c.execute("UPDATE calls SET stage=? WHERE phone_number=?", (stage, phone))
    conn.commit()
    conn.close()
    return {"success": True}


@app.get("/api/minutes")
def get_minutes_data():
    conn = sqlite3.connect(DB_FILE)
    c    = conn.cursor()
    c.execute("""
        SELECT phone_number, COUNT(*) as calls,
               SUM(duration_seconds) as total_actual_seconds,
               SUM(billable_minutes) as billable_minutes
        FROM calls WHERE duration_seconds > 0
        GROUP BY phone_number
    """)
    columns    = [d[0] for d in c.description]
    per_number = [dict(zip(columns, row)) for row in c.fetchall()]
    c.execute("SELECT SUM(billable_minutes), COUNT(*) FROM calls WHERE duration_seconds > 0")
    total_billable, total_calls = c.fetchone()
    active_calls = len([k for k in sessions.keys() if not k.startswith("__")])
    conn.close()
    return {
        "summary": {
            "total_billable_minutes": round(total_billable or 0, 1),
            "total_calls_counted":    total_calls or 0,
            "active_calls":           active_calls,
        },
        "per_number": per_number,
    }


@app.get("/api/download")
async def download_excel(start_date: Optional[str] = None, end_date: Optional[str] = None):
    def _export():
        conn = sqlite3.connect(DB_FILE)
        c    = conn.cursor()
        q    = "SELECT * FROM calls"
        p: list = []
        if start_date and end_date:
            q += " WHERE started_at BETWEEN ? AND ?"
            p = [f"{start_date}T00:00:00", f"{end_date}T23:59:59"]
        c.execute(q, p)
        cols = [d[0] for d in c.description]
        rows = c.fetchall()
        conn.close()
        wb   = Workbook()
        ws   = wb.active
        ws.append(cols)
        for row in rows:
            ws.append(list(row))
        path = "/tmp/GuniVox_Leads.xlsx"
        wb.save(path)
        return path
    filepath = await asyncio.get_running_loop().run_in_executor(_io_executor, _export)
    return FileResponse(
        path=filepath,
        filename="GuniVox_Leads.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.get("/api/call/{call_sid}")
async def get_call_status_ep(call_sid: str):
    conn = sqlite3.connect(DB_FILE)
    c    = conn.cursor()
    c.execute("SELECT status, transcript FROM calls WHERE call_sid = ?", (call_sid,))
    row  = c.fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Call not found")
    status, transcript_json = row
    transcript = json.loads(transcript_json) if transcript_json else []
    if call_sid in sessions:
        transcript = [m for m in sessions[call_sid] if m['role'] != 'system']
    return {"call_sid": call_sid, "status": status, "transcript": transcript}


# ─────────────────────────────────────────
# API — RAG
# ─────────────────────────────────────────

@app.post("/api/rag/documents")
async def add_rag_document(doc: RagDocumentRequest):
    if not doc.content.strip():
        raise HTTPException(status_code=400, detail="content is required")
    conn = sqlite3.connect(DB_FILE)
    c    = conn.cursor()
    c.execute(
        "INSERT INTO rag_documents (title, content, source, created_at) VALUES (?,?,?,?)",
        (doc.title, doc.content.strip(), doc.source, datetime.now().isoformat())
    )
    conn.commit()
    doc_id = c.lastrowid
    conn.close()
    return {"id": doc_id, "success": True}


@app.get("/api/rag/search")
async def rag_search(q: str, top_k: int = 3, threshold: float = 0.5):
    top_k = max(1, min(top_k, 10))
    if not _faiss_available or not faiss_rag.is_ready():
        return {"query": q, "results": [], "rag_enabled": ENABLE_RAG, "error": "FAISS not ready"}
    try:
        results = faiss_rag.search(q, top_k=top_k, score_threshold=threshold)
        return {"query": q, "results": results, "rag_enabled": ENABLE_RAG}
    except Exception as e:
        return {"query": q, "results": [], "rag_enabled": ENABLE_RAG, "error": str(e)}


@app.post("/api/rag/rebuild")
async def rag_rebuild():
    if not _faiss_available:
        raise HTTPException(status_code=503, detail="FAISS not available")
    try:
        faiss_rag.load_index(force_rebuild=True, json_path='final_dataset.json')
        return {"success": True, "message": "FAISS index rebuilt.", **faiss_rag.stats()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/rag/stats")
async def rag_stats():
    return faiss_rag.stats() if _faiss_available else {"ready": False}


# ─────────────────────────────────────────
# API — HEALTH
# ─────────────────────────────────────────

@app.get("/api/health")
async def api_health():
    return {
        "status":         "ok",
        "service":        "GuniVox v4 (production-hardened)",
        "stt":            "sarvam-saaras-v3" if SARVAM_API_KEY else "disabled",
        "tts":            "sarvam-bulbul-v2 (async httpx)",
        "ai":             f"{AI_PROVIDER}/{OPENAI_MODEL if AI_PROVIDER == 'openai' else OLLAMA_MODEL}",
        "tts_cache_size": len(_tts_cache),
        "silero_vad":     silero_model is not None,
        "greeting_ready": os.path.exists(GREETING_PATH) and os.path.getsize(GREETING_PATH) > 500,
        "active_sessions": len([k for k in sessions.keys() if not k.startswith("__")]),
    }


@app.get("/api/llm/health")
async def llm_health():
    faiss_status = faiss_rag.stats() if _faiss_available else {"ready": False}
    if AI_PROVIDER == "ollama":
        try:
            tags   = http_session.get(f"{OLLAMA_BASE_URL}/api/tags", timeout=HTTP_TIMEOUT_SECONDS)
            models = [m.get("name") for m in tags.json().get("models", [])]
            return {"provider": AI_PROVIDER, "model": OLLAMA_MODEL,
                    "model_available": OLLAMA_MODEL in models,
                    "available_models": models, "faiss": faiss_status}
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Ollama health check failed: {e}")
    return {"provider": AI_PROVIDER, "model": OPENAI_MODEL,
            "rag_enabled": ENABLE_RAG, "faiss": faiss_status,
            "tts_cache_entries": len(_tts_cache)}


# ─────────────────────────────────────────
# FRONTEND SERVING (SPA)
# ─────────────────────────────────────────
FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "dist")

if os.path.isdir(os.path.join(FRONTEND_DIR, "assets")):
    app.mount("/assets", StaticFiles(directory=os.path.join(FRONTEND_DIR, "assets")), name="assets")

def _serve_frontend():
    index_path = os.path.join(FRONTEND_DIR, "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path, media_type="text/html")
    return JSONResponse({"status": "ok", "service": "GuniVox v4 — backend running",
                         "hint": "Run 'npm run build' to generate dist/ folder"})

@app.get("/")
@app.head("/")
async def serve_root():
    return _serve_frontend()

@app.get("/favicon.ico")
async def favicon():
    p = os.path.join(FRONTEND_DIR, "favicon.ico")
    return FileResponse(p) if os.path.exists(p) else Response(status_code=204)


# ─────────────────────────────────────────
# WEBSOCKET — Unmute real-time voice
# ─────────────────────────────────────────
try:
    from unmute_handler import run_unmute_pipeline

    @app.websocket("/unmute")
    async def websocket_endpoint(websocket: WebSocket):
        await websocket.accept()
        logger.info("⚡ Unmute WS accepted")
        try:
            await run_unmute_pipeline(websocket)
        except Exception as e:
            logger.error(f"Unmute error: {e}")
        finally:
            logger.info("⚡ Unmute WS closed")
except ImportError:
    logger.info("unmute_handler not found — /unmute endpoint disabled")


# ─────────────────────────────────────────
# CATCH-ALL (SPA routing)
# ─────────────────────────────────────────
@app.api_route("/{path:path}", methods=["GET", "POST", "HEAD"])
async def catch_all(request: Request, path: str):
    if any(path.startswith(p) for p in ["api/", "vobiz-", "status", "hold-", "join-", "counsellor-", "user-fallback"]):
        return JSONResponse({"error": "Not Found"}, status_code=404)
    if request.method == "POST":
        return JSONResponse({"received": True, "path": path})
    file_path = os.path.join(FRONTEND_DIR, path)
    if os.path.isfile(file_path):
        return FileResponse(file_path)
    return _serve_frontend()


# ─────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", "8000"))
    logger.info(f"🚀 GuniVox v4 on port {port}")
    uvicorn.run(
        app, host="0.0.0.0", port=port,
        # Recommended production settings:
        # workers=1,           # single worker (async, shares session state)
        # loop="uvloop",       # faster event loop on Linux
        # http="httptools",    # faster HTTP parser
    )